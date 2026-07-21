from __future__ import annotations

import os
import re
import socket
import threading
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib-cache")

import gradio as gr
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from course_scheduler_agents import (
    CourseWorkflowContext,
    parse_course_spreadsheet,
    run_course_handoff_workflow_sync,
)


load_dotenv()


SECTION_LINE_PATTERN = re.compile(
    r"^\s*(?P<course>.+?)\s*(?:[:：,，]\s*|\s+)(?P<count>\d+)\s*(?:个班|班|分班|sections?|sec)?\s*$",
    re.IGNORECASE,
)
TIME_SLOT_PATTERN = re.compile(
    r"(?:一共|总共|共|设置为|设为)?\s*(?P<count>\d+)\s*(?:个)?\s*(?:time\s*slots?|时间段)\b",
    re.IGNORECASE,
)
SEED_PATTERN = re.compile(r"\bseed\s*[=:：]?\s*(?P<seed>\d+)\b", re.IGNORECASE)
CONFIRM_POSITIVE_PATTERN = re.compile(
    r"(确认|没问题|可以|开始|开跑|跑吧|运行|执行|生成|排吧|按这个|就这样|就按|同意|ok|okay|yes|go)",
    re.IGNORECASE,
)
CONFIRM_NEGATIVE_PATTERN = re.compile(
    r"(不确认|别运行|不要运行|先别|等等|等一下|还不|不是|不对|要改|修改|改一下|重新)",
    re.IGNORECASE,
)

WELCOME_MESSAGE = (
    "您好，欢迎使用智能排课系统。我将协助您核对排课需求，并在确认后生成课程安排。\n\n"
    "请先上传学生选课表，并提供以下信息：各课程的开班数量、排课时段总数、不可安排在同一时段的"
    "课程组合、指定课程的禁排时段，以及各班人数限制。高级设置的默认值为：排课时段数 5、"
    "班级最小人数 12、班级最大人数 30、最大迭代次数 20000。\n\n"
    "信息完整后，系统将汇总全部设置；请您核对并回复“确认运行”，随后正式开始排课。"
)
TYPING_MESSAGE = '<span class="typing-dots"><span></span><span></span><span></span></span>'

DEFAULT_INPUT = ""
OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"

APP_CSS = """
body {
  background: #f7f7f8;
}

.gradio-container {
  max-width: 1040px !important;
  margin: 0 auto !important;
}

#chat_shell {
  min-height: 100vh;
}

#chatbot {
  border: 0;
  background: transparent;
}

#chatbot .message {
  border-radius: 18px !important;
  box-shadow: none !important;
}

#chatbot .user {
  background: #fff4c2 !important;
  color: #1f2937 !important;
  border-radius: 20px 20px 4px 20px !important;
}

#chatbot .bot {
  background: #ffffff !important;
  color: #1f2937 !important;
  border: 1px solid #e5e7eb !important;
  border-radius: 20px 20px 20px 4px !important;
}

#chatbot .bot:has(.system-progress),
#chatbot .message:has(.system-progress),
#chatbot [data-testid="bot"]:has(.system-progress),
#chatbot [data-testid="bot"] .message:has(.system-progress) {
  background: transparent !important;
  border: 0 !important;
  box-shadow: none !important;
  padding: 4px 0 !important;
}

.system-progress {
  width: min(560px, 88%);
  margin: 2px auto;
  color: #6b7280;
  font-size: 12px;
  line-height: 1.35;
}

.system-progress-line {
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 12px;
  margin-bottom: 6px;
}

.system-progress-title {
  color: #374151;
  font-weight: 600;
}

.system-progress-meta {
  white-space: nowrap;
}

.system-progress-track {
  position: relative;
  height: 5px;
  overflow: hidden;
  border-radius: 999px;
  background: #e5e7eb;
}

.system-progress-results {
  display: grid;
  gap: 3px;
  max-height: 180px;
  margin-top: 8px;
  overflow-y: auto;
  color: #4b5563;
  font-variant-numeric: tabular-nums;
}

.system-progress-result {
  display: flex;
  justify-content: space-between;
  gap: 12px;
  padding: 2px 0;
  border-bottom: 1px solid #f3f4f6;
}

.system-progress-bar {
  position: absolute;
  inset: 0 auto 0 0;
  width: 38%;
  border-radius: inherit;
  background: #111827;
  animation: systemProgress 1.25s infinite ease-in-out;
}

@keyframes systemProgress {
  0% {
    transform: translateX(-105%);
  }
  100% {
    transform: translateX(265%);
  }
}

#upload_panel {
  background: transparent !important;
  border: 0 !important;
  box-shadow: none !important;
  padding: 0 !important;
  margin: 8px 0 10px;
  gap: 12px;
}

/* Gradio adds a styled wrapper around Row. Remove that shared outer frame so
   only the two upload cards remain visible. */
.styler:has(> #upload_panel) {
  background: transparent !important;
  border: 0 !important;
  box-shadow: none !important;
  padding: 0 !important;
}

#student_upload_card,
#section_upload_card {
  background: #ffffff !important;
  border: 1px solid #e5e7eb !important;
  border-radius: 16px !important;
  box-shadow: 0 6px 16px rgba(17, 24, 39, 0.05) !important;
  padding: 8px 10px !important;
  min-width: 0;
}

#upload_status,
#section_upload_status {
  color: #4b5563;
  font-size: 13px;
  margin-top: 4px;
}

#student_upload {
  min-height: 0;
  height: auto;
  border: 0 !important;
  box-shadow: none !important;
}

#student_upload label,
#student_upload .wrap,
#student_upload .file-preview,
#student_upload .file-preview-holder,
#student_upload button,
#student_upload [data-testid="file"],
#student_upload [data-testid="file-upload"],
#student_upload .upload-container {
  min-height: 34px !important;
  height: 38px !important;
  padding: 3px 8px !important;
}

#student_upload .wrap.hide {
  display: none !important;
  min-height: 0 !important;
  height: 0 !important;
  padding: 0 !important;
}

#upload_panel .styler,
#upload_panel .block,
#student_upload .file,
#student_upload .file-preview,
#student_upload .file-preview-holder {
  background: transparent !important;
  border: 0 !important;
  box-shadow: none !important;
}

#student_upload .file-preview-holder,
#student_upload .file-preview {
  min-height: 32px !important;
  height: auto !important;
  margin: 0 !important;
  padding: 0 !important;
}

#student_upload table,
#student_upload tbody,
#student_upload tr,
#student_upload td {
  height: 28px !important;
  padding-top: 2px !important;
  padding-bottom: 2px !important;
}

#student_upload .upload-text,
#student_upload .file-preview span,
#student_upload label span {
  font-size: 13px !important;
  line-height: 1.2 !important;
}

#student_upload button {
  max-height: 44px !important;
  overflow: hidden !important;
}

#student_upload .icon-wrap,
#student_upload .or {
  display: none !important;
}

#section_upload {
  min-height: 0;
  height: auto;
  border: 0 !important;
  box-shadow: none !important;
}

#section_upload label,
#section_upload .wrap,
#section_upload .file-preview,
#section_upload .file-preview-holder,
#section_upload button,
#section_upload [data-testid="file"],
#section_upload [data-testid="file-upload"],
#section_upload .upload-container {
  min-height: 34px !important;
  height: 38px !important;
  padding: 3px 8px !important;
}

#section_upload .wrap.hide {
  display: none !important;
  min-height: 0 !important;
  height: 0 !important;
  padding: 0 !important;
}

#section_upload .file,
#section_upload .file-preview,
#section_upload .file-preview-holder {
  background: transparent !important;
  border: 0 !important;
  box-shadow: none !important;
}

#section_upload .file-preview-holder,
#section_upload .file-preview {
  min-height: 32px !important;
  height: auto !important;
  margin: 0 !important;
  padding: 0 !important;
}

#section_upload table,
#section_upload tbody,
#section_upload tr,
#section_upload td {
  height: 28px !important;
  padding-top: 2px !important;
  padding-bottom: 2px !important;
}

#section_upload .upload-text,
#section_upload .file-preview span,
#section_upload label span {
  font-size: 13px !important;
  line-height: 1.2 !important;
}

#section_upload button {
  max-height: 44px !important;
  overflow: hidden !important;
}

#section_upload .icon-wrap,
#section_upload .or {
  display: none !important;
}

#composer {
  background: #ffffff;
  border: 1px solid #e5e7eb;
  border-radius: 24px;
  box-shadow: 0 12px 30px rgba(17, 24, 39, 0.08);
  padding: 12px;
}

#composer textarea {
  border: 0 !important;
  box-shadow: none !important;
  resize: none;
}

#toolbar {
  align-items: center;
  background: transparent !important;
  border: 0 !important;
  box-shadow: none !important;
}

#toolbar *,
#toolbar > div,
#toolbar .block,
#toolbar .form,
#toolbar .gradio-row,
#toolbar .gradio-column {
  background: transparent !important;
  border-color: transparent !important;
  box-shadow: none !important;
}

#send_btn,
#stop_btn {
  min-width: 44px !important;
  width: 44px !important;
  height: 44px !important;
  padding: 0 !important;
  border-radius: 999px !important;
}

#send_btn {
  background: #111111 !important;
  color: #ffffff !important;
  border: 1px solid #111111 !important;
  font-size: 20px !important;
  line-height: 1 !important;
}

#stop_btn {
  background: #9ca3af !important;
  color: #ffffff !important;
  border: 1px solid #9ca3af !important;
  font-size: 15px !important;
  line-height: 1 !important;
}

#advanced_panel,
#results_panel {
  border-radius: 16px !important;
  overflow: hidden;
}

#advanced_panel > button,
#results_panel > button,
#advanced_panel .label-wrap,
#results_panel .label-wrap {
  border-radius: 16px !important;
}

#advanced_panel:focus,
#advanced_panel:focus-visible,
#advanced_panel:focus-within,
#results_panel:focus,
#results_panel:focus-visible,
#results_panel:focus-within,
#advanced_panel > button:focus,
#advanced_panel > button:focus-visible,
#results_panel > button:focus,
#results_panel > button:focus-visible,
#advanced_panel summary:focus,
#advanced_panel summary:focus-visible,
#results_panel summary:focus,
#results_panel summary:focus-visible {
  outline: none !important;
  box-shadow: none !important;
  border-color: #e5e7eb !important;
}

.gradio-container .generating {
  animation: none !important;
  border: 0 !important;
  box-shadow: none !important;
  background: transparent !important;
}

/* Gradio 6.x auto-wraps each Group/Row in a .styler div with a default
   light-gray background (rgb(228, 228, 231) / #e4e4e7). Neutralize it inside
   the panels we custom-style, so the gray fill doesn't bleed through. */
#composer .styler,
#toolbar .styler,
#upload_panel .styler,
#advanced_panel .styler {
  background: transparent !important;
  border: 0 !important;
  box-shadow: none !important;
}

.typing-dots {
  display: inline-flex;
  align-items: center;
  gap: 5px;
  min-width: 34px;
  padding: 3px 0;
}

.typing-dots span {
  width: 7px;
  height: 7px;
  border-radius: 999px;
  background: #9ca3af;
  animation: typingDot 1.1s infinite ease-in-out;
}

.typing-dots span:nth-child(2) {
  animation-delay: 0.15s;
}

.typing-dots span:nth-child(3) {
  animation-delay: 0.3s;
}

@keyframes typingDot {
  0%, 80%, 100% {
    opacity: 0.35;
    transform: translateY(0);
  }
  40% {
    opacity: 1;
    transform: translateY(-3px);
  }
}
"""


def _initial_history() -> List[Dict[str, str]]:
    return [{"role": "assistant", "content": WELCOME_MESSAGE}]


def _initial_memory() -> Dict[str, Any]:
    return {
        "user_messages": [],
        "known_requirements_text": "",
        "section_counts": {},
        "pending_confirmation": False,
        "pending_advanced_override": {},
        "last_settings_summary": "",
        "uploaded_file_path": None,
        "student_count": 0,
        "selected_courses": [],
        "file_warnings": [],
    }


def _chatbot_pairs(messages: List[Dict[str, str]]) -> List[Tuple[Optional[str], Optional[str]]]:
    pairs: List[Tuple[Optional[str], Optional[str]]] = []
    for message in messages:
        if message["role"] == "user":
            pairs.append((message["content"], None))
        elif pairs and pairs[-1][1] is None:
            pairs[-1] = (pairs[-1][0], message["content"])
        else:
            pairs.append((None, message["content"]))
    return pairs


def _message_text(messages: List[Dict[str, str]]) -> str:
    return "\n\n".join(message["content"] for message in messages)


def _user_message_text(messages: List[Dict[str, str]]) -> str:
    return "\n\n".join(
        message["content"] for message in messages if message["role"] == "user"
    )


def _chat_text(messages: List[Dict[str, str]]) -> str:
    return "\n\n".join(f"{message['role']}: {message['content']}" for message in messages)


def _agent_messages(messages: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    agent_messages: List[Dict[str, str]] = []
    for message in messages:
        content = message.get("content")
        if isinstance(content, str):
            text = content
        elif isinstance(content, list):
            text = "\n".join(item for item in content if isinstance(item, str))
        else:
            continue
        if text == TYPING_MESSAGE or "system-progress" in text:
            continue
        agent_messages.append({"role": str(message.get("role", "assistant")), "content": text})
    return agent_messages


def _parse_section_counts(text: str) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for raw_line in text.splitlines():
        match = SECTION_LINE_PATTERN.match(raw_line.strip())
        if match:
            counts[match.group("course").strip()] = int(match.group("count"))
    return counts


def _selected_courses(student_file_path: str) -> List[str]:
    parsed = parse_course_spreadsheet(student_file_path)
    return sorted(
        {
            course
            for course_list in parsed.student_courses.values()
            for course in course_list
            if course and course.strip()
        }
    )


def _coerce_file_path(file_value: Any) -> Optional[str]:
    if not file_value:
        return None
    if isinstance(file_value, str):
        return file_value
    if isinstance(file_value, dict):
        for key in ("path", "name", "orig_name"):
            value = file_value.get(key)
            if isinstance(value, str) and value:
                return value
    if isinstance(file_value, (list, tuple)):
        for item in file_value:
            path = _coerce_file_path(item)
            if path:
                return path
    return None


def _section_counts_for_run(
    messages: List[Dict[str, str]],
    parsed_section_file: Optional[Any] = None,
) -> Dict[str, int]:
    counts = dict(parsed_section_file.section_counts) if parsed_section_file is not None else {}
    counts.update(_parse_section_counts(_user_message_text(messages)))
    return counts


def _missing_section_counts(
    student_file_path: str,
    messages: List[Dict[str, str]],
    parsed_section_file: Optional[Any] = None,
) -> List[str]:
    counts = _section_counts_for_run(messages, parsed_section_file)
    return [course for course in _selected_courses(student_file_path) if course not in counts]


def _is_confirmation_message(text: str) -> bool:
    normalized = text.strip().strip("\"'“”‘’「」")
    if not normalized:
        return False
    if CONFIRM_NEGATIVE_PATTERN.search(normalized):
        return False
    return bool(CONFIRM_POSITIVE_PATTERN.search(normalized))


def _format_section_counts(counts: Dict[str, int]) -> str:
    if not counts:
        return "尚未填写"
    rows = [f"- {course}: {count} 个班" for course, count in sorted(counts.items())]
    if len(rows) > 40:
        rows = rows[:40] + [f"- 还有 {len(counts) - 40} 门课程未显示"]
    return "\n".join(rows)


def _confirmation_message(
    *,
    file_path: str,
    parsed_file: Any,
    messages: List[Dict[str, str]],
    slots: int,
    seed: Optional[float],
    big_course_line: int,
    small_course_line: int,
    class_floor: int,
    class_cap: int,
    max_iterations: int,
    section_counts: Optional[Dict[str, int]] = None,
) -> str:
    selected_courses = sorted(
        {
            course
            for course_list in parsed_file.student_courses.values()
            for course in course_list
            if course and course.strip()
        }
    )
    return (
        "我已经拿到可以排课的信息了。运行前请你确认一下：\n\n"
        f"上传文件：{os.path.basename(file_path)}\n"
        f"学生数量：{len(parsed_file.student_courses)}\n"
        f"课程数量：{len(selected_courses)}\n"
        f"排课时段数：{int(slots)}\n"
        f"班级最小人数：{int(class_floor)}\n"
        f"班级最大人数：{int(class_cap)}\n"
        f"最大迭代次数：{int(max_iterations)}\n\n"
        "课程分班数：\n"
        f"{_format_section_counts(section_counts or _parse_section_counts(_user_message_text(messages)))}\n\n"
        "对话里提到的禁排规则、课程组合和其他偏好均已记录。\n\n"
        "如果这些都对，请回复“确认运行”。如果要改，直接告诉我要改哪一项。"
    )


def _update_memory(
    memory: Optional[Dict[str, Any]],
    messages: List[Dict[str, str]],
    *,
    file_path: Optional[str] = None,
    parsed_file: Optional[Any] = None,
) -> Dict[str, Any]:
    next_memory = dict(memory or _initial_memory())
    if messages:
        user_messages = [
            message["content"]
            for message in messages
            if message.get("role") == "user" and message.get("content")
        ]
        known_text = "\n\n".join(user_messages)

        next_memory["user_messages"] = user_messages
        next_memory["known_requirements_text"] = known_text
        next_memory["section_counts"] = _parse_section_counts(known_text)

    if file_path:
        next_memory["uploaded_file_path"] = file_path

    if parsed_file is not None:
        selected_courses = sorted(
            {
                course
                for course_list in parsed_file.student_courses.values()
                for course in course_list
                if course and course.strip()
            }
        )
        next_memory["student_count"] = len(parsed_file.student_courses)
        next_memory["selected_courses"] = selected_courses
        next_memory["file_warnings"] = list(parsed_file.warnings)

    return next_memory


def _last_int(pattern: re.Pattern[str], text: str, group_name: str) -> Optional[int]:
    matches = list(pattern.finditer(text))
    return int(matches[-1].group(group_name)) if matches else None


def _advanced_conflict_question(messages: List[Dict[str, str]], slots: int, seed: Optional[float]) -> Optional[str]:
    text = _user_message_text(messages)
    mentioned_slots = _last_int(TIME_SLOT_PATTERN, text, "count")
    if mentioned_slots is not None and mentioned_slots != int(slots):
        return f"你文字里写的是 {mentioned_slots} 个 time slots，但高级设置里是 {int(slots)}。要按哪个？"

    mentioned_seed = _last_int(SEED_PATTERN, text, "seed")
    if seed is not None and mentioned_seed is not None and mentioned_seed != int(seed):
        return f"你文字里写 seed={mentioned_seed}，但高级设置里是 {int(seed)}。要按哪个？"

    return None


def _compose_agent_message(
    messages: List[Dict[str, str]],
    slots: int,
    seed: Optional[float],
    big_course_line: int,
    small_course_line: int,
    class_floor: int,
    class_cap: int,
    max_iterations: int,
    memory: Optional[Dict[str, Any]] = None,
) -> str:
    seed_text = "不指定" if seed is None else str(int(seed))
    return (
        "对话历史：\n"
        + _chat_text(_agent_messages(messages))
        + "\n\nSession memory：\n"
        + str(memory or {})
        + "\n\n界面参数：\n"
        + f"- num_time_slots={int(slots)}\n"
        + f"- seed={seed_text}\n"
        + f"- min_students_threshold={int(big_course_line)}\n"
        + f"- max_students_threshold={int(small_course_line)}\n"
        + f"- min_students_per_section={int(class_floor)}\n"
        + f"- max_students_per_section={int(class_cap)}\n"
        + f"- max_iterations={int(max_iterations)}\n\n"
        + "如果信息已经足够，请构造 CourseSchedulingRequest 并调用排课工具。"
    )


def _candidate_metrics(candidates: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "round": candidate.get("round"),
            "candidate": candidate.get("candidate"),
            "seed": candidate.get("seed"),
            "accepted": candidate.get("accepted"),
            **candidate.get("result", {}).get("metrics", {}),
        }
        for candidate in candidates
    ]


def _schedule_table(result: Dict[str, Any]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for slot, sections in sorted(result.get("schedule_by_slot", {}).items(), key=lambda item: int(item[0])):
        for section in sections:
            course = section.rsplit("_", 1)[0] if "_" in section else section
            rows.append([int(slot) + 1, section, course])
    return rows


def _schedule_workbook_path(result: Dict[str, Any]) -> str:
    schedule_by_slot = result.get("schedule_by_slot") or {}
    if not schedule_by_slot:
        raise ValueError("Schedule result does not include schedule_by_slot.")

    ordered_slots = sorted(schedule_by_slot.items(), key=lambda item: int(item[0]))
    max_sections = max((len(sections) for _, sections in ordered_slots), default=0)
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    output_path = OUTPUT_DIR / f"course_schedule_{timestamp}.xlsx"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Class Schedule"
    sheet.sheet_view.showGridLines = False

    title_range = f"A1:{get_column_letter(len(ordered_slots))}1"
    sheet.merge_cells(title_range)
    title = sheet["A1"]
    title.value = "Class Schedule"
    title.font = Font(bold=True, size=16, color="1F2937")
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.fill = PatternFill("solid", fgColor="F8FAFC")
    sheet.row_dimensions[1].height = 30

    header_fill = PatternFill("solid", fgColor="E5E7EB")
    header_font = Font(bold=True, color="111827")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    cell_fill = PatternFill("solid", fgColor="FFFFFF")

    for col_index, (slot, _) in enumerate(ordered_slots, start=1):
        cell = sheet.cell(row=2, column=col_index, value=f"Time Slot {int(slot) + 1}")
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col_index)].width = 28

    for row_offset in range(max_sections):
        row_index = row_offset + 3
        sheet.row_dimensions[row_index].height = 28
        for col_index, (_, sections) in enumerate(ordered_slots, start=1):
            value = sections[row_offset] if row_offset < len(sections) else ""
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.fill = cell_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.freeze_panes = "A3"
    workbook.save(output_path)
    return str(output_path)


def _summary(run_result: Any) -> str:
    best = run_result.best_schedule.get("result", {})
    metrics = best.get("metrics", {})
    return "\n".join(
        [
            run_result.schedule_summary,
            "",
            f"验收通过: {'是' if run_result.accepted else '否'}",
            f"排课执行次数: {len(run_result.candidates)}",
            f"重试轮数: {run_result.retry_rounds_used}",
            f"满意率: {metrics.get('satisfaction_rate')}%",
            f"冲突学生数: {metrics.get('conflict_count')}",
            f"总成本: {metrics.get('total_cost')}",
        ]
    )


def _progress_stage(stage: str) -> Tuple[str, str]:
    stages = {
        "conversation": ("启动任务", "正在核对并转交排课需求"),
        "orchestrator": ("整理需求", "正在提取并校验排课约束"),
        "scheduler": ("准备排课", "已接收结构化排课需求"),
        "tool": ("生成候选", "排课工具正在搜索可用方案"),
        "summary": ("整理结果", "正在汇总候选方案指标"),
        "completed": ("课表生成完成", "正在展示结果"),
    }
    return stages.get(stage, ("处理请求", "排课流程正在运行"))


def _system_progress_message(
    elapsed_seconds: float,
    *,
    stage: str,
    done: bool = False,
    run_progress: Optional[List[Dict[str, Any]]] = None,
) -> str:
    elapsed = max(0, int(elapsed_seconds))
    if done:
        title = "课表生成完成"
        detail = "正在整理结果"
        meta = f"{elapsed} 秒"
        bar = ""
    else:
        title, detail = _progress_stage(stage)
        meta = f"{elapsed} 秒"
        bar = '<div class="system-progress-track"><div class="system-progress-bar"></div></div>'

    result_rows = ""
    for item in run_progress or []:
        satisfaction = item.get("satisfaction_rate")
        satisfaction_text = (
            "—"
            if satisfaction is None
            else f"{float(satisfaction):.2f}%"
        )
        conflicts = item.get("conflict_count")
        conflict_text = "—" if conflicts is None else str(int(conflicts))
        result_rows += (
            '<div class="system-progress-result">'
            f'<span>第 {int(item.get("run", 0))} 次</span>'
            f'<span>满意度 {satisfaction_text} · 冲突 {conflict_text} 人</span>'
            '</div>'
        )
    results = (
        f'<div class="system-progress-results">{result_rows}</div>'
        if result_rows
        else ""
    )

    return (
        '<div class="system-progress">'
        '<div class="system-progress-line">'
        f'<span><span class="system-progress-title">{title}</span> · {detail}</span>'
        f'<span class="system-progress-meta">{meta}</span>'
        '</div>'
        f'{bar}'
        f'{results}'
        '</div>'
    )


def _ask_missing_sections(missing: List[str]) -> str:
    shown = "\n".join(f"- {course}" for course in missing[:20])
    suffix = "" if len(missing) <= 20 else f"\n还有 {len(missing) - 20} 门没有列出。"
    return f"我还缺这些课的分班数，先补一下再排：\n{shown}{suffix}"


def _conversation_context(
    *,
    issue: str,
    slots: int,
    seed: Optional[float],
    big_course_line: int,
    small_course_line: int,
    class_floor: int,
    class_cap: int,
    max_iterations: int,
    file_path: Optional[str] = None,
    parsed_file: Optional[Any] = None,
    missing: Optional[List[str]] = None,
    conflict_question: Optional[str] = None,
    error: Optional[str] = None,
    memory: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    selected_courses: List[str] = []
    student_count = 0
    file_warnings: List[str] = []
    if parsed_file is not None:
        student_count = len(parsed_file.student_courses)
        selected_courses = sorted(
            {
                course
                for course_list in parsed_file.student_courses.values()
                for course in course_list
                if course and course.strip()
            }
        )
        file_warnings = list(parsed_file.warnings)

    return {
        "issue": issue,
        "uploaded_file_path": file_path,
        "uploaded_file_has_student_courses": bool(student_count),
        "student_count": student_count,
        "selected_courses": selected_courses[:80],
        "missing_section_counts": (missing or [])[:80],
        "conflict_question": conflict_question,
        "error": error,
        "advanced_settings": {
            "time_slots": int(slots),
            "seed": None if seed is None else int(seed),
            "班级最小人数": int(class_floor),
            "班级最大人数": int(class_cap),
            "max_iterations": int(max_iterations),
        },
        "file_warnings": file_warnings,
        "session_memory": memory or _initial_memory(),
        "requires_final_confirmation_before_running": True,
        "instruction_to_agent": (
            "Reply to the user conversationally. Ask for missing scheduling details before running. "
            "When information is complete, the app will summarize every setting and require user confirmation."
        ),
    }


def _fallback_agent_reply(
    *,
    issue: str,
    missing: Optional[List[str]] = None,
    conflict_question: Optional[str] = None,
    error: Optional[str] = None,
) -> str:
    if issue == "missing_file":
        return (
            "我可以先帮你整理排课需求，不过真正生成课表需要学生选课表。"
            "你先上传包含学生和课程列的 CSV/Excel 文件，然后告诉我每门课开几个班、"
            "time slots、课程冲突/禁排 slot，以及班级最小和最大人数限制。"
        )
    if issue == "file_read_error":
        return f"我没法读取这个上传文件：{error}"
    if issue == "invalid_student_file":
        return "这个文件里没有识别到学生选课数据。请上传包含学生和课程列的 CSV/Excel 文件。"
    if issue == "advanced_conflict":
        return conflict_question or "你的文字需求和高级设置有冲突，我需要先确认按哪个为准。"
    if issue == "missing_section_counts":
        return _ask_missing_sections(missing or [])
    if issue == "missing_api_key":
        return "我可以继续对话，但要真正调用排课 agent，需要先在 .env 里填写 OPENAI_API_KEY。"
    return "我需要再确认一点信息，才能继续排课。"


def _agent_reply(
    history: List[Dict[str, str]],
    context: Dict[str, Any],
    *,
    model: str,
    fallback: str,
) -> str:
    if not os.getenv("OPENAI_API_KEY"):
        return fallback
    try:
        result = run_course_handoff_workflow_sync(
            _agent_messages(history),
            context,
            model=model,
            ready_to_schedule=False,
        )
        return result.message
    except Exception as exc:
        return f"{fallback}\n\n对话 agent 暂时没能生成回复：{exc}"


def _append_reply(
    history: List[Dict[str, str]],
    reply: str,
) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], Dict[str, Any], str, List[List[Any]], Dict[str, Any], Dict[str, Any], List[Dict[str, Any]], Any]:
    history.append({"role": "assistant", "content": reply})
    return history, history, _initial_memory(), "", [], {}, {}, [], gr.update(value=""), gr.update(visible=False, value=None)


def _output(
    history: List[Dict[str, str]],
    *,
    memory: Optional[Dict[str, Any]] = None,
    summary: str = "",
    schedule_table: Optional[List[List[Any]]] = None,
    metrics: Optional[Dict[str, Any]] = None,
    structured_request: Optional[Dict[str, Any]] = None,
    candidate_metrics: Optional[List[Dict[str, Any]]] = None,
    slots_update: Optional[Any] = None,
    seed_update: Optional[Any] = None,
) -> Tuple[Any, ...]:
    return (
        history,
        history,
        memory or _initial_memory(),
        summary,
        schedule_table or [],
        metrics or {},
        structured_request or {},
        candidate_metrics or [],
        gr.update(value=""),
        slots_update or gr.update(),
        seed_update or gr.update(),
    )


def _set_status(history: List[Dict[str, str]], status: str) -> None:
    if history and history[-1]["role"] == "assistant":
        history[-1]["content"] = status
    else:
        history.append({"role": "assistant", "content": status})


def _append_assistant(history: List[Dict[str, str]], content: str) -> None:
    history.append({"role": "assistant", "content": content})


def handle_file_upload(
    student_file_path: Optional[str],
    memory_state: Optional[Dict[str, Any]],
) -> Tuple[str, Dict[str, Any]]:
    file_path = _coerce_file_path(student_file_path)
    if not file_path:
        return "尚未上传学生选课表。", dict(memory_state or _initial_memory())

    filename = os.path.basename(file_path)

    try:
        parsed_file = parse_course_spreadsheet(file_path)
    except Exception as exc:
        memory = _update_memory(memory_state, [], file_path=file_path)
        return f"已上传 `{filename}`，但暂时读不了：{exc}", memory

    memory = _update_memory(
        memory_state,
        [],
        file_path=file_path,
        parsed_file=parsed_file,
    )
    course_count = len(memory.get("selected_courses", []))
    if parsed_file.student_courses:
        return (
            f"已上传 `{filename}`。识别到 {len(parsed_file.student_courses)} 名学生、{course_count} 门课程。",
            memory,
        )

    warning_text = "；".join(parsed_file.warnings)
    suffix = f"\n解析提示：{warning_text}" if warning_text else ""
    return (
        f"已上传 `{filename}`，但没有识别到学生选课数据。请确认文件里有学生和课程列。{suffix}",
        memory,
    )


def handle_section_upload(section_file_path: Optional[str]) -> str:
    file_path = _coerce_file_path(section_file_path)
    if not file_path:
        return "尚未上传课程分班表。"
    filename = os.path.basename(file_path)
    try:
        parsed_file = parse_course_spreadsheet(file_path)
    except Exception as exc:
        return f"已上传 `{filename}`，但暂时读不了：{exc}"
    if not parsed_file.section_counts:
        return f"已上传 `{filename}`，但没有识别到课程分班数。"
    return f"已上传 `{filename}`。识别到 {len(parsed_file.section_counts)} 门课程的分班数。"


def _show_running_buttons() -> Tuple[Any, Any]:
    return gr.update(visible=False), gr.update(visible=True)


def _show_idle_buttons() -> Tuple[Any, Any]:
    return gr.update(visible=True), gr.update(visible=False)


def _stop_running(
    chat_history: Optional[List[Dict[str, str]]],
) -> Tuple[Any, Any, Any, Any]:
    history = list(chat_history or _initial_history())
    stopped_message = "当前操作已终止。"
    if history and history[-1].get("role") == "assistant":
        content = history[-1].get("content")
        if content == TYPING_MESSAGE or (
            isinstance(content, str) and "system-progress" in content
        ):
            history[-1]["content"] = stopped_message
        elif content != stopped_message:
            history.append({"role": "assistant", "content": stopped_message})
    else:
        history.append({"role": "assistant", "content": stopped_message})
    return (
        history,
        history,
        gr.update(visible=True),
        gr.update(visible=False),
    )


def respond(
    user_message: str,
    chat_history: Optional[List[Dict[str, str]]],
    memory_state: Optional[Dict[str, Any]],
    student_file_path: Optional[str],
    section_file_path: Optional[str],
    slots: int,
    seed: Optional[float],
    max_iterations: int,
    big_course_line: int,
    small_course_line: int,
    class_floor: int,
    class_cap: int,
):
    history = list(chat_history or _initial_history())
    if user_message and user_message.strip():
        history.append({"role": "user", "content": user_message.strip()})

    model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
    memory = _update_memory(memory_state, history)
    _append_assistant(history, TYPING_MESSAGE)
    yield _output(history, memory=memory)

    file_path = _coerce_file_path(student_file_path)
    if not file_path:
        reply = _agent_reply(
            history,
            _conversation_context(
                issue="missing_file",
                slots=slots,
                seed=seed,
                big_course_line=big_course_line,
                small_course_line=small_course_line,
                class_floor=class_floor,
                class_cap=class_cap,
                max_iterations=max_iterations,
                memory=memory,
            ),
            model=model,
            fallback=_fallback_agent_reply(issue="missing_file"),
        )
        _set_status(history, reply)
        memory = _update_memory(memory, history)
        yield _output(history, memory=memory)
        return

    _set_status(history, TYPING_MESSAGE)
    memory = _update_memory(memory, history, file_path=file_path)
    yield _output(history, memory=memory)

    try:
        parsed_file = parse_course_spreadsheet(file_path)
    except Exception as exc:
        reply = _agent_reply(
            history,
            _conversation_context(
                issue="file_read_error",
                slots=slots,
                seed=seed,
                big_course_line=big_course_line,
                small_course_line=small_course_line,
                class_floor=class_floor,
                class_cap=class_cap,
                max_iterations=max_iterations,
                file_path=file_path,
                error=str(exc),
                memory=memory,
            ),
            model=model,
            fallback=_fallback_agent_reply(issue="file_read_error", error=str(exc)),
        )
        _set_status(history, reply)
        memory = _update_memory(memory, history, file_path=file_path)
        yield _output(history, memory=memory)
        return

    parsed_section_file = None
    section_path = _coerce_file_path(section_file_path)
    if section_path:
        try:
            parsed_section_file = parse_course_spreadsheet(section_path)
        except Exception as exc:
            _set_status(history, f"课程分班表读取失败：{exc}")
            yield _output(history, memory=memory)
            return
        if not parsed_section_file.section_counts:
            _set_status(history, "课程分班表里没有识别到 `course` 和 `section_count` 数据。")
            yield _output(history, memory=memory)
            return

    memory = _update_memory(memory, history, file_path=file_path, parsed_file=parsed_file)

    if not parsed_file.student_courses:
        warning_text = "；".join(parsed_file.warnings)
        fallback = _fallback_agent_reply(issue="invalid_student_file")
        if warning_text:
            fallback += f"\n解析提示：{warning_text}"
        reply = _agent_reply(
            history,
            _conversation_context(
                issue="invalid_student_file",
                slots=slots,
                seed=seed,
                big_course_line=big_course_line,
                small_course_line=small_course_line,
                class_floor=class_floor,
                class_cap=class_cap,
                max_iterations=max_iterations,
                file_path=file_path,
                parsed_file=parsed_file,
                memory=memory,
            ),
            model=model,
            fallback=fallback,
        )
        _set_status(history, reply)
        memory = _update_memory(memory, history, file_path=file_path, parsed_file=parsed_file)
        yield _output(history, memory=memory)
        return

    latest_user_text = (user_message or "").strip()
    pending_override = dict(memory.get("pending_advanced_override") or {})
    if pending_override:
        explicit_slots = _last_int(TIME_SLOT_PATTERN, latest_user_text, "count")
        explicit_seed = _last_int(SEED_PATTERN, latest_user_text, "seed")
        confirmed = _is_confirmation_message(latest_user_text)
        next_slots = explicit_slots
        next_seed = explicit_seed
        if confirmed:
            if next_slots is None:
                next_slots = pending_override.get("slots")
            if next_seed is None:
                next_seed = pending_override.get("seed")

        if next_slots is not None or next_seed is not None:
            slots_update = gr.update()
            seed_update = gr.update()
            if next_slots is not None:
                slots = int(next_slots)
                slots_update = gr.update(value=slots)
            if next_seed is not None:
                seed = int(next_seed)
                seed_update = gr.update(value=seed)
            memory["pending_advanced_override"] = {}
            yield _output(
                history,
                memory=memory,
                slots_update=slots_update,
                seed_update=seed_update,
            )

    conflict_question = _advanced_conflict_question(history, int(slots), seed)
    if conflict_question:
        text = _user_message_text(history)
        mentioned_slots = _last_int(TIME_SLOT_PATTERN, text, "count")
        mentioned_seed = _last_int(SEED_PATTERN, text, "seed")
        pending_override = {}
        if mentioned_slots is not None and mentioned_slots != int(slots):
            pending_override["slots"] = mentioned_slots
        if seed is not None and mentioned_seed is not None and mentioned_seed != int(seed):
            pending_override["seed"] = mentioned_seed
        memory["pending_advanced_override"] = pending_override
        reply = _agent_reply(
            history,
            _conversation_context(
                issue="advanced_conflict",
                slots=slots,
                seed=seed,
                big_course_line=big_course_line,
                small_course_line=small_course_line,
                class_floor=class_floor,
                class_cap=class_cap,
                max_iterations=max_iterations,
                file_path=file_path,
                parsed_file=parsed_file,
                conflict_question=conflict_question,
                memory=memory,
            ),
            model=model,
            fallback=_fallback_agent_reply(
                issue="advanced_conflict",
                conflict_question=conflict_question,
            ),
        )
        _set_status(history, reply)
        memory = _update_memory(memory, history, file_path=file_path, parsed_file=parsed_file)
        yield _output(history, memory=memory)
        return

    section_counts = _section_counts_for_run(history, parsed_section_file)
    missing = _missing_section_counts(file_path, history, parsed_section_file)
    if missing:
        memory["pending_confirmation"] = False
        reply = _agent_reply(
            history,
            _conversation_context(
                issue="missing_section_counts",
                slots=slots,
                seed=seed,
                big_course_line=big_course_line,
                small_course_line=small_course_line,
                class_floor=class_floor,
                class_cap=class_cap,
                max_iterations=max_iterations,
                file_path=file_path,
                parsed_file=parsed_file,
                missing=missing,
                memory=memory,
            ),
            model=model,
            fallback=_fallback_agent_reply(issue="missing_section_counts", missing=missing),
        )
        _set_status(history, reply)
        memory = _update_memory(memory, history, file_path=file_path, parsed_file=parsed_file)
        yield _output(history, memory=memory)
        return

    if memory.get("pending_confirmation") and _is_confirmation_message(latest_user_text):
        memory["pending_confirmation"] = False
    else:
        confirmation = _confirmation_message(
            file_path=file_path,
            parsed_file=parsed_file,
            messages=history,
            slots=slots,
            seed=seed,
            big_course_line=big_course_line,
            small_course_line=small_course_line,
            class_floor=class_floor,
            class_cap=class_cap,
            max_iterations=max_iterations,
            section_counts=section_counts,
        )
        memory["pending_confirmation"] = True
        memory["last_settings_summary"] = confirmation
        confirmation_context = _conversation_context(
            issue="confirmation_required",
            slots=slots,
            seed=seed,
            big_course_line=big_course_line,
            small_course_line=small_course_line,
            class_floor=class_floor,
            class_cap=class_cap,
            max_iterations=max_iterations,
            file_path=file_path,
            parsed_file=parsed_file,
            memory=memory,
        )
        confirmation_context["settings_summary"] = confirmation
        reply = _agent_reply(
            history,
            confirmation_context,
            model=model,
            fallback=confirmation,
        )
        _set_status(history, reply)
        memory = _update_memory(memory, history, file_path=file_path, parsed_file=parsed_file)
        memory["pending_confirmation"] = True
        memory["last_settings_summary"] = confirmation
        yield _output(history, memory=memory)
        return

    if not os.getenv("OPENAI_API_KEY"):
        _set_status(history, _fallback_agent_reply(issue="missing_api_key"))
        memory = _update_memory(memory, history, file_path=file_path, parsed_file=parsed_file)
        yield _output(history, memory=memory)
        return

    request_overrides = {
        "num_time_slots": int(slots),
        "seed": None if seed is None else int(seed),
        "min_students_threshold": int(big_course_line),
        "max_students_threshold": int(small_course_line),
        "min_students_per_section": int(class_floor),
        "max_students_per_section": int(class_cap),
        "max_iterations": int(max_iterations),
        "candidate_runs": 2,
        "include_schedule": True,
        "include_section_loads": True,
    }
    workflow_context = CourseWorkflowContext(
        student_courses=dict(parsed_file.student_courses),
        section_counts=section_counts,
        request_overrides=request_overrides,
        parsed_spreadsheets=[
            item for item in (parsed_file, parsed_section_file) if item is not None
        ],
    )
    scheduler_state: Dict[str, Any] = {
        "done": False,
        "result": None,
        "error": None,
        "workflow_context": workflow_context,
    }

    def _run_scheduler() -> None:
        try:
            workflow_result = run_course_handoff_workflow_sync(
                _agent_messages(history),
                _conversation_context(
                    issue="ready_to_schedule",
                    slots=slots,
                    seed=seed,
                    big_course_line=big_course_line,
                    small_course_line=small_course_line,
                    class_floor=class_floor,
                    class_cap=class_cap,
                    max_iterations=max_iterations,
                    file_path=file_path,
                    parsed_file=parsed_file,
                    memory=memory,
                ),
                model=model,
                workflow_context=workflow_context,
                ready_to_schedule=True,
            )
            if workflow_result.schedule is None:
                raise RuntimeError("Handoff workflow returned no schedule.")
            scheduler_state["result"] = workflow_result.schedule
        except Exception as exc:
            scheduler_state["error"] = exc
        finally:
            scheduler_state["done"] = True

    started_at = time.monotonic()
    worker = threading.Thread(target=_run_scheduler, daemon=True)
    worker.start()
    displayed_seconds = 0
    displayed_stage = workflow_context.stage
    displayed_run_count = 0
    try:
        _set_status(
            history,
            _system_progress_message(0, stage=workflow_context.stage),
        )
        yield _output(history, memory=memory)
        while not scheduler_state["done"] or (time.monotonic() - started_at) < 1.2:
            time.sleep(0.2)
            elapsed_seconds = int(time.monotonic() - started_at)
            current_stage = workflow_context.stage
            current_progress = list(workflow_context.run_progress)
            current_run_count = len(current_progress)
            if (
                elapsed_seconds == displayed_seconds
                and current_stage == displayed_stage
                and current_run_count == displayed_run_count
            ):
                continue
            displayed_seconds = elapsed_seconds
            displayed_stage = current_stage
            displayed_run_count = current_run_count
            _set_status(
                history,
                _system_progress_message(
                    displayed_seconds,
                    stage=current_stage,
                    run_progress=current_progress,
                ),
            )
            yield _output(history, memory=memory)
    finally:
        if not scheduler_state["done"]:
            workflow_context.cancel_requested = True
            print("[cancel] Scheduling cancellation requested", flush=True)

    worker.join(timeout=0)

    if scheduler_state["error"] is not None:
        exc = scheduler_state["error"]
        _set_status(history, f"运行时出错：{exc}")
        memory = _update_memory(memory, history, file_path=file_path, parsed_file=parsed_file)
        yield _output(history, memory=memory)
        return

    _set_status(
        history,
        _system_progress_message(
            time.monotonic() - started_at,
            stage=workflow_context.stage,
            done=True,
            run_progress=list(workflow_context.run_progress),
        ),
    )
    yield _output(history, memory=memory)

    run_result = scheduler_state["result"]
    best = run_result.best_schedule.get("result", {})
    schedule_file_path = _schedule_workbook_path(best)
    assistant_message = _summary(run_result)
    schedule_attachment = {
        "path": schedule_file_path,
        "orig_name": os.path.basename(schedule_file_path),
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "meta": {"_type": "gradio.FileData"},
    }
    history.append(
        {
            "role": "assistant",
            "content": [assistant_message, schedule_attachment],
        }
    )
    memory = _update_memory(memory, history, file_path=file_path, parsed_file=parsed_file)
    yield _output(
        history,
        memory=memory,
        summary=assistant_message,
        schedule_table=_schedule_table(best),
        metrics=best.get("metrics", {}),
        structured_request=run_result.structured_request.model_dump(),
        candidate_metrics=_candidate_metrics(run_result.candidates),
    )


def reset_chat() -> Tuple[Any, ...]:
    history = _initial_history()
    return (
        history,
        history,
        _initial_memory(),
        "尚未上传学生选课表。",
        "尚未上传课程分班表。",
        "",
        [],
        {},
        {},
        [],
        gr.update(value=None),
        gr.update(value=None),
    )


def build_app() -> gr.Blocks:
    with gr.Blocks(title="Course Arrangement Agent") as demo:
        with gr.Column(elem_id="chat_shell"):
            chatbot = gr.Chatbot(
                value=_initial_history(),
                label=None,
                height=520,
                elem_id="chatbot",
                sanitize_html=False,
            )
            chat_state = gr.State(_initial_history())
            memory_state = gr.State(_initial_memory())
            big_course_line = gr.State(25)
            small_course_line = gr.State(65)

            with gr.Row(elem_id="upload_panel"):
                with gr.Group(elem_id="student_upload_card"):
                    with gr.Column():
                        student_file = gr.File(
                            label="学生选课 CSV",
                            file_types=[".csv", ".tsv", ".xlsx", ".xlsm", ".xls"],
                            type="filepath",
                            elem_id="student_upload",
                        )
                        upload_status = gr.Markdown(
                            "尚未上传学生选课表。",
                            elem_id="upload_status",
                        )
                with gr.Group(elem_id="section_upload_card"):
                    with gr.Column():
                        section_file = gr.File(
                            label="课程分班 CSV",
                            file_types=[".csv", ".tsv", ".xlsx", ".xlsm", ".xls"],
                            type="filepath",
                            elem_id="section_upload",
                        )
                        section_upload_status = gr.Markdown(
                            "尚未上传课程分班表。",
                            elem_id="section_upload_status",
                        )

            with gr.Group(elem_id="composer"):
                user_box = gr.Textbox(
                    value=DEFAULT_INPUT,
                    lines=3,
                    max_lines=8,
                    label=None,
                    placeholder="告诉我你的排课需求，例如：AP Chemistry 和 AP Biology 不要同 slot...",
                    show_label=False,
                )
                with gr.Row(elem_id="toolbar"):
                    with gr.Column(scale=12):
                        gr.Markdown("")
                    with gr.Column(scale=1, min_width=58):
                        send_button = gr.Button("↑", elem_id="send_btn")
                        stop_button = gr.Button(
                            "■",
                            variant="stop",
                            visible=False,
                            elem_id="stop_btn",
                        )

            with gr.Accordion("高级设置", open=False, elem_id="advanced_panel"):
                with gr.Row():
                    slots = gr.Slider(2, 12, value=5, step=1, label="排课时段数")
                    seed = gr.Number(value=None, precision=0, visible=False)
                    max_iterations = gr.Number(
                        value=20000,
                        precision=0,
                        minimum=100,
                        maximum=100000,
                        step=100,
                        label="最大迭代次数",
                    )
                with gr.Row():
                    class_floor = gr.Number(value=12, precision=0, label="班级最小人数")
                    class_cap = gr.Number(value=30, precision=0, label="班级最大人数")

            with gr.Accordion("结果详情", open=False, elem_id="results_panel"):
                summary = gr.Textbox(label="总结", lines=8)
                schedule_table = gr.Dataframe(
                    headers=["slot", "section", "course"],
                    label="课表",
                    interactive=False,
                )
                with gr.Row():
                    metrics = gr.JSON(label="Metrics")
                    structured_request = gr.JSON(label="结构化输入")
                candidate_metrics = gr.JSON(label="每轮结果")

            clear_button = gr.Button("清空对话", variant="secondary")

        inputs = [
            user_box,
            chat_state,
            memory_state,
            student_file,
            section_file,
            slots,
            seed,
            max_iterations,
            big_course_line,
            small_course_line,
            class_floor,
            class_cap,
        ]
        outputs = [
            chatbot,
            chat_state,
            memory_state,
            summary,
            schedule_table,
            metrics,
            structured_request,
            candidate_metrics,
            user_box,
            slots,
            seed,
        ]
        send_prepare = send_button.click(
            _show_running_buttons,
            inputs=[],
            outputs=[send_button, stop_button],
            queue=False,
        )
        send_event = send_prepare.then(respond, inputs=inputs, outputs=outputs)
        send_event.then(
            _show_idle_buttons,
            inputs=[],
            outputs=[send_button, stop_button],
            queue=False,
        )

        submit_prepare = user_box.submit(
            _show_running_buttons,
            inputs=[],
            outputs=[send_button, stop_button],
            queue=False,
        )
        submit_event = submit_prepare.then(respond, inputs=inputs, outputs=outputs)
        submit_event.then(
            _show_idle_buttons,
            inputs=[],
            outputs=[send_button, stop_button],
            queue=False,
        )
        stop_button.click(
            _stop_running,
            inputs=[chat_state],
            outputs=[chatbot, chat_state, send_button, stop_button],
            cancels=[send_event, submit_event],
            queue=False,
        )
        student_file.change(
            handle_file_upload,
            inputs=[student_file, memory_state],
            outputs=[upload_status, memory_state],
        )
        section_file.change(
            handle_section_upload,
            inputs=[section_file],
            outputs=[section_upload_status],
        )
        clear_button.click(
            reset_chat,
            inputs=[],
            outputs=[
                chatbot,
                chat_state,
                memory_state,
                upload_status,
                section_upload_status,
                summary,
                schedule_table,
                metrics,
                structured_request,
                candidate_metrics,
                student_file,
                section_file,
            ],
        )
    return demo


def _find_available_port(start: int = 7860, end: int = 8060) -> int:
    for port in range(start, end + 1):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            try:
                sock.bind(("127.0.0.1", port))
            except OSError:
                continue
            return port
    raise RuntimeError(f"No available local port in range {start}-{end}.")


demo = build_app()


if __name__ == "__main__":
    launch_options: Dict[str, Any] = {"css": APP_CSS}
    deployment_port = os.getenv("PORT")
    if deployment_port:
        launch_options.update(
            server_name="0.0.0.0",
            server_port=int(deployment_port),
        )
    elif not os.getenv("SPACE_ID"):
        launch_options.update(
            server_name="127.0.0.1",
            server_port=_find_available_port(),
        )
    demo.launch(**launch_options)
